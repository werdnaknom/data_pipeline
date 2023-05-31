import typing

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import xlsxwriter

from io import BytesIO

from .formatter import Formatter


class XLSXFormatter(Formatter):

    def _format(self, output, sheets: typing.Dict[str, pd.DataFrame]):
        writer = pd.ExcelWriter(output, engine="xlsxwriter")
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()
        return output

    def _create_excel_document(self, input: typing.Tuple[str, pd.DataFrame]):
        # TODO:: NOT USED
        filename = input[0]
        dataframe = input[1]
        # Create a new workbook
        workbook = xlsxwriter.Workbook(filename)

        # Define a format for the header row
        header_format = workbook.add_format({'bold': True, 'align': 'center'})

        # Iterate over the processed dataframes
        for index, dataframe in enumerate(dataframes):
            # Create a new worksheet in the workbook
            sheet_name = f"Sheet{index + 1}"
            worksheet = workbook.add_worksheet(sheet_name)

            # Write the dataframe to the worksheet
            header_row = dataframe.columns.tolist()
            data_rows = dataframe.values.tolist()

            worksheet.write_row(0, 0, header_row, header_format)
            worksheet.add_table(1, 0, len(data_rows), len(header_row) - 1, {'data': data_rows})

            # Apply formatting to the sheet
            self.apply_sheet_formatting(worksheet)

        # Close the workbook
        workbook.close()

    def apply_sheet_formatting(self, worksheet):
        # TODO:: NOT USED
        # Apply formatting to the data rows
        cell_format = worksheet.get_default_row()
        cell_format.set_align('left')
        worksheet.set_default_row(cell_format=cell_format)


class OpenpyxlFormatter(Formatter):
    # TODO:: NOT USED

    def format(self, input: typing.Dict[str, pd.DataFrame]):
        # TODO:: NOT USED
        pass

    def create_excel_document(dataframes, filename):
        # TODO:: NOT USED
        # Create a new workbook
        workbook = Workbook()

        # Iterate over the processed dataframes
        for index, dataframe in enumerate(dataframes):
            # Create a new sheet in the workbook
            sheet_name = f"Sheet{index + 1}"
            sheet = workbook.create_sheet(title=sheet_name)

            # Write the dataframe to the sheet
            for row in dataframe_to_rows(dataframe, index=False, header=True):
                sheet.append(row)

            # Apply formatting to the sheet
            apply_sheet_formatting(sheet)

        # Remove the default sheet created by openpyxl
        workbook.remove(workbook["Sheet"])

        # Save the workbook as an Excel file
        workbook.save(filename)

    def apply_sheet_formatting(self, sheet):
        # TODO:: NOT USED
        # Apply formatting to the header row
        header_row = sheet[1]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Apply formatting to the data rows
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
